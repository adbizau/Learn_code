from openpyxl import Workbook
from openpyxl import load_workbook

workbook = load_workbook("D:\Moz\All_test.xlsx")
workbook.sheetnames
sheet = workbook.active
print(len(sheet["A"]))